from openpyxl import Workbook, load_workbook

taille_grille=(8,8)

repertoire='snake/'
nom_fichier='donnees'
suffixe='.xlsx'

score_max=taille_grille[0]*taille_grille[1]
taille_grille=str(taille_grille)





fichier=load_workbook(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)

def creer_feuille(nom_feuille):
    fichier.create_sheet(nom_feuille)

    fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)

def supprimer_feuille(nom_feuille):
    fichier.remove_sheet(fichier[nom_feuille])

    fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)

def renommer_feuille(nom_feuille,nouveau_nom_feuille):
    feuille=fichier[nom_feuille]
    feuille.title=nouveau_nom_feuille

    fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)

def nettoyer_feuille(nom_feuille):
    feuille=fichier[nom_feuille]
    feuille.delete_rows(1,feuille.max_row)

    fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)

def initialiser_feuille(nom_feuille):
    feuille=fichier[nom_feuille]
    feuille['A1']='score'
    feuille['B1']='nombre pas moyen'
    feuille['C1']='nombre essais'
    feuille['D1']='somme'
    feuille['E1']='nombre parties perdues'
    feuille['F1']='proportion parties_perdues/parties_jouees'

    for i in range(score_max):
        feuille['A'+str(i+2)]=i
        feuille['B'+str(i+2)]=0
        feuille['C'+str(i+2)]=0
        feuille['D'+str(i+2)]=0
        feuille['E'+str(i+2)]=0
        feuille['F'+str(i+2)]=0
    
    fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)

def existence_feuille(nom_feuille):
    return nom_feuille in fichier.sheetnames

'''il faut mettre en commentaire l'import du fichier en tant que variable fichier (au debut) si fichier corrompu'''
def recreer_fichier(liste_noms_feuilles):
    global fichier
    fichier=Workbook()
    renommer_feuille('Sheet',liste_noms_feuilles[0])
    initialiser_feuille(liste_noms_feuilles[0])
    for i in range(1,len(liste_noms_feuilles)):
        creer_feuille(liste_noms_feuilles[i])
        initialiser_feuille(liste_noms_feuilles[i])

def nettoyer_fichier():
    for nom_feuille in fichier.sheetnames:
        nettoyer_feuille(nom_feuille)
        initialiser_feuille(nom_feuille)


def ajouter_liste_score(nom_feuille,liste_score,etat_partie):
    feuille=fichier[nom_feuille]
    for i in range(score_max):

        if i<len(liste_score):

            # on ajoute à la somme
            feuille['D'+str(i+2)]=feuille['D'+str(i+2)].value+liste_score[i]

            # on ajoute au nombre d'essais
            feuille['C'+str(i+2)]=feuille['C'+str(i+2)].value+1

            # on recalcule les moyennes
            feuille['B'+str(i+2)]=feuille['D'+str(i+2)].value/feuille['C'+str(i+2)].value

        
        
        # on met a jour les proportions
        feuille['F'+str(i+2)]=feuille['E'+str(i+2)].value/feuille['C2'].value
    
    if etat_partie==-1:
        feuille['E'+str(len(liste_score)+2)]=feuille['E'+str(len(liste_score)+2)].value+1
        
    
    fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)


'''instructions ici'''
# creer_feuille('a_etoile_graphe_oriente_avec_cout')
# nettoyer_feuille('raccourcis_croissants_v1')
# nettoyer_fichier()
# creer_feuille('a_etoile_graphe_oriente_hami')

nettoyer_feuille('a_etoile_seulement_si_modif')
initialiser_feuille('a_etoile_seulement_si_modif')










# fichier.save(repertoire+nom_fichier+taille_grille.replace(' ', '')+suffixe)
# fichier.close()



# faire qu'on puisse faire des acquisitions de donnees sur toutes les tailles facilement
# faire backups
# faire qu'on a le nombre de parties perdus en fonction du score
