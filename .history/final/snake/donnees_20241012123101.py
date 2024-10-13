from openpyxl import Workbook, load_workbook

class Donnees:
    def __init__(self,nb_cases):
        self.taille_grille=(nb_cases,nb_cases)
    
    def initialiser(self):
        

        self.repertoire='final/snake/'

        

        self.suffixe='.xlsx'

        self.score_max=self.taille_grille[0]*self.taille_grille[1]
        self.taille_grille=str(self.taille_grille).replace(' ', '')





        

        self.a_executer()

    def creer_feuille(self,nom_feuille):
        self.fichier.create_sheet(nom_feuille)

        # self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille.replace(' ', '')+self.suffixe)
        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille+self.suffixe)

    def supprimer_feuille(self,nom_feuille):
        self.fichier.remove_sheet(self.fichier[nom_feuille])

        # self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille.replace(' ', '')+self.suffixe)
        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille+self.suffixe)

    def renommer_feuille(self,nom_feuille,nouveau_nom_feuille):
        feuille=self.fichier[nom_feuille]
        feuille.title=nouveau_nom_feuille

        # self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille.replace(' ', '')+self.suffixe)
        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille+self.suffixe)

    def nettoyer_feuille(self,nom_feuille):
        feuille=self.fichier[nom_feuille]
        feuille.delete_rows(1,feuille.max_row)

        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille.replace(' ', '')+self.suffixe)

    def initialiser_feuille(self,nom_feuille):
        feuille=self.fichier[nom_feuille]
        feuille['A1']='score'
        feuille['B1']='nombre pas moyen'
        feuille['C1']='nombre essais'
        feuille['D1']='somme'
        feuille['E1']='nombre parties perdues'
        feuille['F1']='proportion parties_perdues/parties_jouees'

        for i in range(self.score_max):
            feuille['A'+str(i+2)]=i
            feuille['B'+str(i+2)]=0
            feuille['C'+str(i+2)]=0
            feuille['D'+str(i+2)]=0
            feuille['E'+str(i+2)]=0
            feuille['F'+str(i+2)]=0
        
        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille.replace(' ', '')+self.suffixe)

    def existence_feuille(self,nom_feuille):
        return nom_feuille in self.fichier.sheetnames

    '''il faut mettre en commentaire l'import du fichier en tant que variable fichier (au debut) si fichier corrompu'''
    def recreer_fichier(self,liste_noms_feuilles):
        global fichier
        fichier=Workbook()
        self.renommer_feuille('Sheet',liste_noms_feuilles[0])
        self.initialiser_feuille(liste_noms_feuilles[0])
        for i in range(1,len(liste_noms_feuilles)):
            self.creer_feuille(liste_noms_feuilles[i])
            self.initialiser_feuille(liste_noms_feuilles[i])

    def nettoyer_fichier(self):
        for nom_feuille in self.fichier.sheetnames:
            self.nettoyer_feuille(nom_feuille)
            self.initialiser_feuille(nom_feuille)


    def ajouter_liste_score(self,nom_feuille,liste_score,etat_partie):
        feuille=self.fichier[nom_feuille]
        for i in range(self.score_max):

            if i<len(liste_score):

                # on ajoute à la somme
                feuille['D'+str(i+2)]=feuille['D'+str(i+2)].value+liste_score[i]

                # on ajoute au nombre d'essais
                feuille['C'+str(i+2)]=feuille['C'+str(i+2)].value+1

                # on recalcule les moyennes
                feuille['B'+str(i+2)]=feuille['D'+str(i+2)].value/feuille['C'+str(i+2)].value

            # on met a jour les proportions
            feuille['F'+str(i+2)]=feuille['E'+str(i+2)].value/feuille['C2'].value
        
        if etat_partie==-1:
            feuille['E'+str(len(liste_score)+2)]=feuille['E'+str(len(liste_score)+2)].value+1
            
        
        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille.replace(' ', '')+self.suffixe)
    
    def a_executer(self):

        """init"""
        # self.nom_fichier='donnees'
        # self.nom_fichier='donnees_raccourcis_risque'

        self.fichier=load_workbook(self.repertoire+self.nom_fichier+self.taille_grille+self.suffixe)


        '''instructions ici'''
        # self.creer_feuille('raccourcis_croissants_v3')
        # self.supprimer_feuille('donnees_raccourcis_risque1')
        # nettoyer_feuille('raccourcis_croissants_v1')
        # self.nettoyer_fichier()
        # creer_feuille('a_etoile_graphe_oriente_hami')


        # self.nettoyer_feuille('raccourcis_croissants_v3')
        # self.initialiser_feuille('raccourcis_croissants_v3')









        self.fichier.save(self.repertoire+self.nom_fichier+self.taille_grille+self.suffixe)
        

if __name__=='__main__':
    donnees=Donnees()
    donnees.nom_fichier='donnees_raccourcis_risque'
    donnees.initialiser()
    donnees.a_executer()

# faire qu'on puisse faire des acquisitions de donnees sur toutes les tailles facilement
# faire backups
# faire qu'on a le nombre de parties perdus en fonction du score
